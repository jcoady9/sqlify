import openpyxl

import sqlite3

def create(db_name, spreadsheet_file):
    workbook = openpyxl.load_workbook(spreadsheet_file)
    worksheet = workbook.active

    conn = sqlite3.connect(db_name)
    cursor = conn.cursor()

    create_cmd = 'CREATE TABLE IF NOT EXISTS {} '.format('table_name')

    columns = list()
    for col_num in range(1, worksheet.max_column + 1):
        header = worksheet.cell(row=1,column=col_num)
        col_name = header.value.strip().replace(' ','_')
        cell = worksheet.cell(row=2,column=col_num)

        if cell.data_type == cell.TYPE_NUMERIC:
            columns.append('{} INTEGER'.format(col_name))
        elif cell.data_type == cell.TYPE_STRING:
            columns.append('{} TEXT'.format(col_name))
        elif cell.is_date:
            columns.append('{} TEXT'.format(col_name))
        else:
            raise NotImplementedError('Datatype not supported.')

    create_cmd += '({})'.format(','.join(col for col in columns))
    print(create_cmd)

    cursor.execute(create_cmd)
    conn.commit()

    #Populate database
    rows = list()
    for row_num in range(2, worksheet.max_row + 1):
        row = list()
        for col_num in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row=row_num,column=col_num)
            row.append(cell.value)
        print(row)
        rows.append(tuple(row))

    insert_cmd = 'INSERT INTO {} VALUES ({})'.format('table_name', ','.join(['?' for x in range(0,worksheet.max_column)]))

    cursor.executemany(insert_cmd, rows)
    conn.commit()

    conn.close()

def insert(db_name, spreadsheet_file):
    conn = sqlite3.connect(db_name)
    cursor = conn.cursor()
    workbook = openpyxl.load_workbook(spreadsheet_file)
    worksheet = workbook.active

    rows = list()
    for row_num in range(2, worksheet.max_row + 1):
        row = list()
        for col_num in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row=row_num,column=col_num)
            row.append(cell.value)
        print(row)
        rows.append(tuple(row))

    insert_cmd = 'INSERT INTO {} VALUES ({})'.format('table_name', ','.join(['?' for x in range(0,worksheet.max_column)]))

    cursor.executemany(insert_cmd, rows)
    conn.commit()
    conn.close()

#TODO: UPDATE
def update(db_name, spreadsheet_file, update_on):
    conn = sqlite3.connect(db_name)
    cursor = conn.cursor()
    workbook = openpyxl.load_workbook(spreadsheet_file)
    worksheet = workbook.active

    worksheet_rows = list()
    for row_num in range(2, worksheet.max_row + 1):
        row = list()
        for col_num in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row=row_num,column=col_num)
            row.append(cell.value)
        print(row)
        worksheet_rows.append(tuple(row))

    select_cmd = "SELECT * FROM {}".format('table_name')
    db_rows = cursor.execute(select_cmd)


    updatable_rows = list()
    for row in db_rows:
        ws_row = worksheet_rows.pop()
        for i in range(len(row)):
            if row[i] == ws_row[i]:
                updatable_rows.append(ws_row)
                break

    update_cmd = 'UPDATE VALUES {} IN {} WHERE {} = {}'.format(','.join(['?' for x in range(0,worksheet.max_column)], 'table_name', update_on, 15))

    cursor.executemany(update_cmd, updateble_rows)
    conn.close()

#TODO: MAIN

create('test_db.sqlite3', 'test.xlsx')
insert('test_db.sqlite3', 'test2.xlsx')
update('test_db.sqlite3', 'test3.xlsx', 15)
