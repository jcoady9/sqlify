"""
This script allows the user to create a table in a SQLite database using data from a spreadsheet (*.xls, *.xlsx).
The script will create a new table using the first row of the spreadsheet as column names and the remaining row will be 
entered into the table as data.
"""

from openpyxl import load_workbook

from datetime import datetime
import argparse
import sqlite3
import os

def slugify(text):
    """Replaces any preceding or trailing whitespace from a given string and replaces all spaces with an underscore"""
    return text.strip().replace(' ', '_')

def get_sql_data_type(cell):
    """Determine the data type of the a given cell in a worksheet."""
    if cell.data_type == cell.TYPE_NUMERIC:
        return 'INTEGER'
    elif cell.data_type == cell.TYPE_STRING:
        return 'TEXT'
    elif cell.is_date:
        return 'TEXT'
    else:
        raise NotImplementedError('Datatype for cell {} is not supported'.format(cell.coordinate))

def get_columns(worksheet):
    """
    Gets the values stored in the cells of the first row of the worksheet. Returns a list of sequence objects containing
    the name of the columns and the data type for the column i.e. [( column_name, sql_data_type), ... ]
    """
    db_columns = list()
    for column in worksheet.iter_cols(min_col=1, max_row=2):
        db_columns.append((slugify(column[0].value), get_sql_data_type(column[1])))
    return db_columns

def get_values(worksheet, start_row=1):
    """
    Get the values stored in the cells from the second row onwards. Returns a list of each row of the worksheet as a list
    i.e. [ [row2_col1_value, row2_col2_value, ... ], [row3_col1_value, ... ], ... ]. Cells that contain a datetime object
    as the value will be converted into string object with the 'mm/dd/yy' format.
    """
    values = list()
    for row in worksheet.iter_rows(min_row=2):
        db_row = list()
        for cell in row:
            if cell.is_date:
                db_row.append(cell.value.strftime('%x'))
            elif cell.data_type == cell.TYPE_STRING:
                db_row.append(cell.value.strip().replace('\'', '\'\''))
            else:
                db_row.append(cell.value)
        values.append(db_row)
    return values

def create_table(conn, cursor, table_name, columns):
    """Creates and executes CREATE TABLE statement using given table name and columns."""
    column_declarations = ','.join(['{} {}'.format(column[0], column[1]) for column in columns])
    create_stmt = 'CREATE TABLE IF NOT EXISTS {}({})'.format(table_name, column_declarations)
    cursor.execute(create_stmt)
    conn.commit()

def insert_values(conn, cursor, table_name, rows):
    """Create and executes INSERT VALUES statement using the given table name and rows."""
    column_names = ','.join(rows.pop(0))
    values = [tuple(row) for row in rows]
    insert_stmt = 'INSERT INTO {}({}) VALUES ({})'.format(table_name, column_names, ','.join(['?' for x in range(0, len(values[0]))]))
    cursor.executemany(insert_stmt, values)
    conn.commit()

def main():
    parser = argparse.ArgumentParser(description='A simple python script to create a sqlite database from a xlsx spreadsheet file.')
    parser.add_argument('--db', type=str, default='', help='Path to the SQLite database. If no path is provided, a new database named [file] will be created.')
    parser.add_argument('file', type=str, help='Path of the Excel file.')
    args = parser.parse_args()

    workbook = load_workbook(args.file)
    worksheet = workbook.active

    table_name = slugify(worksheet.title)

    db_name = args.db or '.'.join([os.path.splitext(os.path.basename(args.file))[0], 'sqlite3'])

    conn = sqlite3.connect(db_name)
    cursor = conn.cursor()

    # take first row, use as column names
    db_columns = get_columns(worksheet)

    create_table(conn, cursor, table_name, db_columns)

    # prep data for inserting into new table
    values = [[column[0] for column in db_columns]]
    values.extend(get_values(worksheet, start_row=2))

    insert_values(conn, cursor, table_name, values)
        
if __name__ == '__main__':
    main()
